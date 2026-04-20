import math
import re
import threading
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import requests
from openpyxl import load_workbook

try:
    import searoute as sr
except Exception:
    sr = None

APP_TITLE = "HLI Project GHG Calcs Builder"
USER_AGENT = "HLI-Project-GHG-Calcs/3.0 (local desktop app)"
REQUEST_TIMEOUT = 14

INPUT_HEADERS = [
    'Project ID', 'Pick Up Date', 'Weight (Pounds)', 'Origin', 'Destination', 'Mode',
    'Delivery Date', 'Distance (Miles)', 'Subcontractor/Partner', 'Client',
    '(Road) Vehicle category (US)', '(Sea) Vessel Type', '(Inland Water) Vessel Type'
]
COLS_A_TO_N = list(range(1, 15))

ALIASES = {
    'loredo, tx': 'Laredo, TX',
    'loredo tx': 'Laredo, TX',
    'new york city, ny': 'New York, NY',
    'port of freeport': 'Port Freeport, Freeport, TX',
    'port of houston': 'Port of Houston, Houston, TX',
    'los angeles, california': 'Los Angeles, CA',
    'cartagena, colombia': 'Cartagena, Colombia',
    'oakland, ca': 'Oakland, CA',
    'oakland, california': 'Oakland, CA',
    'cancun, mexico': 'Cancun, Quintana Roo, Mexico',
    'boston, ma': 'Boston, MA',
    'ithica, ny': 'Ithaca, NY',
    'ithaca, ny': 'Ithaca, NY',
    'myrtle beach, sc': 'Myrtle Beach, SC',
    'miami, florida': 'Miami, FL',
    'rotterdam, netherlands': 'Rotterdam, Netherlands',
}

PORT_HINTS = {
    'the hague, netherlands': 'Port of Rotterdam, Netherlands',
    'rotterdam, netherlands': 'Port of Rotterdam, Netherlands',
    'new york, ny': 'Port of New York, NY',
    'new york city, ny': 'Port of New York, NY',
    'los angeles, california': 'Port of Los Angeles, CA',
    'los angeles, ca': 'Port of Los Angeles, CA',
    'cartagena, colombia': 'Port of Cartagena, Colombia',
    'port of freeport': 'Port Freeport, Freeport, TX',
    'port of houston': 'Port of Houston, Houston, TX',
    'myrtle beach, sc': 'Port of Charleston, SC',
    'miami, fl': 'PortMiami, Miami, FL',
    'miami, florida': 'PortMiami, Miami, FL',
}

CITY_COORDS = {
    'laredo, tx': (27.5306, -99.4803),
    'laredo, texas': (27.5306, -99.4803),
    'los angeles, ca': (34.0522, -118.2437),
    'los angeles, california': (34.0522, -118.2437),
    'cartagena, colombia': (10.3910, -75.4794),
    'new york, ny': (40.7128, -74.0060),
    'new york city, ny': (40.7128, -74.0060),
    'houston, tx': (29.7604, -95.3698),
    'freeport, tx': (28.9541, -95.3597),
    'charleston, sc': (32.7765, -79.9311),
    'rotterdam, netherlands': (51.9244, 4.4777),
    'the hague, netherlands': (52.0705, 4.3007),
    'oakland, ca': (37.8044, -122.2711),
    'oakland, california': (37.8044, -122.2711),
    'cancun, quintana roo, mexico': (21.1619, -86.8515),
    'cancun, mexico': (21.1619, -86.8515),
    'boston, ma': (42.3601, -71.0589),
    'boston, massachusetts': (42.3601, -71.0589),
    'ithaca, ny': (42.4430, -76.5019),
    'ithica, ny': (42.4430, -76.5019),
    'myrtle beach, sc': (33.6891, -78.8867),
    'miami, fl': (25.7617, -80.1918),
    'miami, florida': (25.7617, -80.1918),
    'port of los angeles, ca': (33.7361, -118.2626),
    'port of cartagena, colombia': (10.3997, -75.5144),
    'port of houston, houston, tx': (29.7284, -95.2650),
    'port freeport, freeport, tx': (28.9365, -95.3088),
    'port of rotterdam, netherlands': (51.8850, 4.2867),
    'port of new york, ny': (40.6840, -74.0419),
    'portmiami, miami, fl': (25.7781, -80.1794),
    'port of charleston, sc': (32.7812, -79.9361),
}

STATE_COORDS = {
    'al': (32.8067, -86.7911), 'alabama': (32.8067, -86.7911),
    'ak': (61.3707, -152.4044), 'alaska': (61.3707, -152.4044),
    'az': (33.7298, -111.4312), 'arizona': (33.7298, -111.4312),
    'ar': (34.9697, -92.3731), 'arkansas': (34.9697, -92.3731),
    'ca': (36.1162, -119.6816), 'california': (36.1162, -119.6816),
    'co': (39.0598, -105.3111), 'colorado': (39.0598, -105.3111),
    'ct': (41.5978, -72.7554), 'connecticut': (41.5978, -72.7554),
    'de': (39.3185, -75.5071), 'delaware': (39.3185, -75.5071),
    'fl': (27.7663, -81.6868), 'florida': (27.7663, -81.6868),
    'ga': (33.0406, -83.6431), 'georgia': (33.0406, -83.6431),
    'hi': (21.0943, -157.4983), 'hawaii': (21.0943, -157.4983),
    'id': (44.2405, -114.4788), 'idaho': (44.2405, -114.4788),
    'il': (40.3495, -88.9861), 'illinois': (40.3495, -88.9861),
    'in': (39.8494, -86.2583), 'indiana': (39.8494, -86.2583),
    'ia': (42.0115, -93.2105), 'iowa': (42.0115, -93.2105),
    'ks': (38.5266, -96.7265), 'kansas': (38.5266, -96.7265),
    'ky': (37.6681, -84.6701), 'kentucky': (37.6681, -84.6701),
    'la': (31.1695, -91.8678), 'louisiana': (31.1695, -91.8678),
    'ma': (42.2302, -71.5301), 'massachusetts': (42.2302, -71.5301),
    'md': (39.0639, -76.8021), 'maryland': (39.0639, -76.8021),
    'me': (44.6939, -69.3819), 'maine': (44.6939, -69.3819),
    'mi': (43.3266, -84.5361), 'michigan': (43.3266, -84.5361),
    'mn': (45.6945, -93.9002), 'minnesota': (45.6945, -93.9002),
    'ms': (32.7416, -89.6787), 'mississippi': (32.7416, -89.6787),
    'mo': (38.4561, -92.2884), 'missouri': (38.4561, -92.2884),
    'nc': (35.6301, -79.8064), 'north carolina': (35.6301, -79.8064),
    'nj': (40.2989, -74.5210), 'new jersey': (40.2989, -74.5210),
    'ny': (42.1657, -74.9481), 'new york': (42.1657, -74.9481),
    'oh': (40.3888, -82.7649), 'ohio': (40.3888, -82.7649),
    'pa': (40.5908, -77.2098), 'pennsylvania': (40.5908, -77.2098),
    'ri': (41.6809, -71.5118), 'rhode island': (41.6809, -71.5118),
    'sc': (33.8569, -80.9450), 'south carolina': (33.8569, -80.9450),
    'tn': (35.7478, -86.6923), 'tennessee': (35.7478, -86.6923),
    'tx': (31.0545, -97.5635), 'texas': (31.0545, -97.5635),
    'va': (37.7693, -78.1700), 'virginia': (37.7693, -78.1700),
}

COUNTRY_COORDS = {
    'united states': (39.8283, -98.5795), 'usa': (39.8283, -98.5795), 'us': (39.8283, -98.5795),
    'mexico': (23.6345, -102.5528),
    'canada': (56.1304, -106.3468),
    'colombia': (4.5709, -74.2973),
    'netherlands': (52.1326, 5.2913),
}

ROAD_FACTOR = 1.33
RAIL_FACTOR = 1.42
INLAND_FACTOR = 1.28
SEA_FACTOR = 1.08

METERS_PER_MILE = 1609.344
KM_PER_MILE = 1.609344

ATLANTIC_ICW_MILES = {
    'norfolk, va': 0,
    'wilmington, nc': 295,
    'myrtle beach, sc': 350,
    'charleston, sc': 455,
    'savannah, ga': 565,
    'jacksonville, fl': 760,
    'west palm beach, fl': 1010,
    'miami, fl': 1090,
}

RAIL_HUBS = {
    'boston, ma': (42.3601, -71.0589),
    'albany, ny': (42.6526, -73.7562),
    'selkirk, ny': (42.5656, -73.7985),
    'syracuse, ny': (43.0481, -76.1474),
    'binghamton, ny': (42.0987, -75.9180),
    'ithaca, ny': (42.4430, -76.5019),
    'chicago, il': (41.8781, -87.6298),
    'kansas city, mo': (39.0997, -94.5786),
    'memphis, tn': (35.1495, -90.0490),
    'new orleans, la': (29.9511, -90.0715),
    'dallas, tx': (32.7767, -96.7970),
    'el paso, tx': (31.7619, -106.4850),
    'los angeles, ca': (34.0522, -118.2437),
    'oakland, ca': (37.8044, -122.2711),
    'miami, fl': (25.7617, -80.1918),
    'atlanta, ga': (33.7490, -84.3880),
    'jacksonville, fl': (30.3322, -81.6557),
}
RAIL_CORRIDOR_EDGES = {
    ('boston, ma', 'albany, ny'): 205,
    ('albany, ny', 'selkirk, ny'): 15,
    ('selkirk, ny', 'syracuse, ny'): 145,
    ('selkirk, ny', 'binghamton, ny'): 135,
    ('binghamton, ny', 'ithaca, ny'): 75,
    ('syracuse, ny', 'ithaca, ny'): 70,
    ('jacksonville, fl', 'miami, fl'): 365,
    ('atlanta, ga', 'jacksonville, fl'): 350,
    ('new orleans, la', 'dallas, tx'): 530,
    ('dallas, tx', 'el paso, tx'): 635,
    ('el paso, tx', 'los angeles, ca'): 800,
    ('los angeles, ca', 'oakland, ca'): 405,
    ('memphis, tn', 'atlanta, ga'): 385,
    ('chicago, il', 'memphis, tn'): 530,
    ('chicago, il', 'kansas city, mo'): 510,
    ('kansas city, mo', 'dallas, tx'): 520,
}

_geocode_cache = {}
_route_cache = {}
_distance_cache = {}


def miles_from_meters(value):
    return float(value) / METERS_PER_MILE


def miles_from_km(value):
    return float(value) / KM_PER_MILE


def round_miles(value):
    if value is None:
        return None
    return max(1, round(float(value)))


def ensure_final_miles(value):
    """Final guard before writing to Excel.

    All estimator functions are expected to return miles already. This helper
    simply normalizes type/rounding so the workbook always receives numeric miles.
    """
    if value is None:
        return None
    return round_miles(value)


def normalize_text(value):
    if value is None:
        return ''
    text = str(value).strip()
    lowered = re.sub(r'\s+', ' ', text.lower())
    return ALIASES.get(lowered, text)


def normalize_key(value):
    return re.sub(r'\s+', ' ', normalize_text(value).strip().lower())


def looks_blank_or_zero(value):
    if value is None:
        return True
    if isinstance(value, str):
        text = value.strip()
        return text in {'', '0', '0.0', 'NAME?', '#NAME?'}
    return value == 0


def clean_output_value(value):
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if text in {'', '0', '0.0', 'NAME?', '#NAME?'}:
            return None
        return text
    if value == 0:
        return None
    return value


def parse_headers(ws, header_row=1):
    mapping = {}
    for c in range(1, ws.max_column + 1):
        header = ws.cell(header_row, c).value
        if header is not None:
            mapping[str(header).strip()] = c
    return mapping


def _make_geo_result(lat, lon, label, source):
    return {'lat': float(lat), 'lon': float(lon), 'display_name': label, 'source': source}


def _candidate_texts(query):
    q = normalize_text(query)
    candidates = []
    if q:
        candidates.append(q)
    low = normalize_key(q)
    if low in PORT_HINTS:
        candidates.append(PORT_HINTS[low])
    compact = q.replace('Port of ', '').replace('Port ', '')
    if compact and compact not in candidates:
        candidates.append(compact)
    for cand in list(candidates):
        parts = [p.strip() for p in cand.split(',') if p.strip()]
        if len(parts) >= 2:
            simplified = ', '.join(parts[:2])
            if simplified not in candidates:
                candidates.append(simplified)
    return candidates


def _lookup_builtin(query):
    for cand in _candidate_texts(query):
        key = normalize_key(cand)
        if key in CITY_COORDS:
            lat, lon = CITY_COORDS[key]
            return _make_geo_result(lat, lon, cand, 'builtin-city')
        parts = [p.strip().lower() for p in cand.split(',') if p.strip()]
        for part in reversed(parts):
            if part in STATE_COORDS:
                lat, lon = STATE_COORDS[part]
                return _make_geo_result(lat, lon, cand, 'builtin-state')
            if part in COUNTRY_COORDS:
                lat, lon = COUNTRY_COORDS[part]
                return _make_geo_result(lat, lon, cand, 'builtin-country')
    return None


def geocode_location(query):
    query = normalize_text(query)
    if not query:
        return None
    if query in _geocode_cache:
        return _geocode_cache[query]

    builtin = _lookup_builtin(query)
    if builtin:
        _geocode_cache[query] = builtin
        return builtin

    for candidate in _candidate_texts(query):
        try:
            response = requests.get(
                'https://nominatim.openstreetmap.org/search',
                params={'format': 'jsonv2', 'limit': 1, 'q': candidate},
                headers={'User-Agent': USER_AGENT},
                timeout=REQUEST_TIMEOUT,
            )
            response.raise_for_status()
            data = response.json()
            if data:
                result = {
                    'lat': float(data[0]['lat']),
                    'lon': float(data[0]['lon']),
                    'display_name': data[0]['display_name'],
                    'source': 'online-geocode',
                }
                _geocode_cache[query] = result
                return result
        except Exception:
            continue

    builtin = _lookup_builtin(query)
    _geocode_cache[query] = builtin
    return builtin


def haversine_miles(a, b):
    lat1, lon1 = math.radians(a['lat']), math.radians(a['lon'])
    lat2, lon2 = math.radians(b['lat']), math.radians(b['lon'])
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    h = math.sin(dlat / 2) ** 2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon / 2) ** 2
    return 3958.7613 * 2 * math.asin(math.sqrt(h))


def road_distance_miles(origin, destination):
    origin = normalize_text(origin)
    destination = normalize_text(destination)
    key = ('road', origin, destination)
    if key in _route_cache:
        return _route_cache[key]
    a = geocode_location(origin)
    b = geocode_location(destination)
    if not a or not b:
        _route_cache[key] = None
        return None
    try:
        response = requests.get(
            f'https://router.project-osrm.org/route/v1/driving/{a["lon"]},{a["lat"]};{b["lon"]},{b["lat"]}',
            params={'overview': 'false'},
            headers={'User-Agent': USER_AGENT},
            timeout=REQUEST_TIMEOUT,
        )
        response.raise_for_status()
        data = response.json()
        routes = data.get('routes') or []
        if not routes:
            _route_cache[key] = None
            return None
        miles = miles_from_meters(routes[0]['distance'])
        gc = haversine_miles(a, b)
        miles = max(miles, gc * 1.12, gc * ROAD_FACTOR if gc > 1000 else miles)
        _route_cache[key] = miles
        return miles
    except Exception:
        _route_cache[key] = None
        return None


def _nearest_icw_point(label):
    key = normalize_key(label)
    if key in ATLANTIC_ICW_MILES:
        return key
    geo = geocode_location(label)
    if not geo:
        return None
    best = None
    best_dist = float('inf')
    for name in ATLANTIC_ICW_MILES:
        lat, lon = CITY_COORDS.get(name, (None, None))
        if lat is None:
            continue
        d = haversine_miles(geo, {'lat': lat, 'lon': lon})
        if d < best_dist:
            best = name
            best_dist = d
    return best if best_dist <= 120 else None


def inland_water_distance_miles(origin, destination):
    key = ('inland', normalize_text(origin), normalize_text(destination))
    if key in _route_cache:
        return _route_cache[key]
    o_icw = _nearest_icw_point(origin)
    d_icw = _nearest_icw_point(destination)
    if o_icw and d_icw:
        miles = abs(ATLANTIC_ICW_MILES[d_icw] - ATLANTIC_ICW_MILES[o_icw])
        if miles > 0:
            _route_cache[key] = (miles, f'Atlantic ICW corridor estimate ({o_icw.title()} ↔ {d_icw.title()})')
            return _route_cache[key]
    a = geocode_location(origin)
    b = geocode_location(destination)
    if a and b:
        gc = haversine_miles(a, b)
        miles = round_miles(gc * INLAND_FACTOR)
        result = (miles, 'Inland water estimate (great-circle × calibrated corridor factor)')
        _route_cache[key] = result
        return result
    return None


def sea_distance_miles(origin, destination):
    key = ('sea', normalize_text(origin), normalize_text(destination))
    if key in _route_cache:
        return _route_cache[key]

    origin_q = PORT_HINTS.get(normalize_key(origin), normalize_text(origin))
    destination_q = PORT_HINTS.get(normalize_key(destination), normalize_text(destination))
    a = geocode_location(origin_q)
    b = geocode_location(destination_q)
    if not a or not b:
        return None

    if sr is not None:
        try:
            route = sr.searoute([a['lon'], a['lat']], [b['lon'], b['lat']], units='km', include_ports=True)
            length_km = route.properties.get('length')
            if length_km:
                miles = round_miles(miles_from_km(length_km))
                method = 'Sea estimate (maritime network route)'
                _route_cache[key] = (miles, method)
                return _route_cache[key]
        except Exception:
            pass

    gc = haversine_miles(a, b)
    miles = round_miles(gc * SEA_FACTOR)
    result = (miles, 'Sea estimate (great-circle × calibrated sea factor)')
    _route_cache[key] = result
    return result


def _nearest_rail_hub(label):
    geo = geocode_location(label)
    if not geo:
        return None
    best = None
    best_dist = float('inf')
    for name, (lat, lon) in RAIL_HUBS.items():
        d = haversine_miles(geo, {'lat': lat, 'lon': lon})
        if d < best_dist:
            best = name
            best_dist = d
    return (best, best_dist) if best is not None else None


def _rail_graph_distance(start, end):
    # simple Dijkstra on corridor edges
    graph = {}
    for (u, v), w in RAIL_CORRIDOR_EDGES.items():
        graph.setdefault(u, []).append((v, w))
        graph.setdefault(v, []).append((u, w))
    import heapq
    pq = [(0, start)]
    seen = set()
    while pq:
        dist, node = heapq.heappop(pq)
        if node == end:
            return dist
        if node in seen:
            continue
        seen.add(node)
        for nxt, w in graph.get(node, []):
            if nxt not in seen:
                heapq.heappush(pq, (dist + w, nxt))
    return None


def rail_distance_miles(origin, destination):
    key = ('rail', normalize_text(origin), normalize_text(destination))
    if key in _route_cache:
        return _route_cache[key]
    a = geocode_location(origin)
    b = geocode_location(destination)
    if not a or not b:
        return None
    gc = haversine_miles(a, b)
    o_hub = _nearest_rail_hub(origin)
    d_hub = _nearest_rail_hub(destination)
    if o_hub and d_hub:
        (o_name, o_last), (d_name, d_last) = o_hub, d_hub
        corridor = _rail_graph_distance(o_name, d_name)
        if corridor is not None:
            miles = round_miles(corridor + o_last * 1.15 + d_last * 1.15)
            miles = max(miles, round(gc * 1.18))
            result = (miles, f'Rail estimate (network corridor via {o_name.title()} → {d_name.title()})')
            _route_cache[key] = result
            return result
    road = road_distance_miles(origin, destination)
    if road is not None:
        miles = max(round_miles(road * 1.12), round_miles(gc * RAIL_FACTOR))
        result = (miles, 'Rail estimate (road network proxy × rail factor)')
        _route_cache[key] = result
        return result
    miles = round_miles(gc * RAIL_FACTOR)
    result = (miles, 'Rail estimate (great-circle × calibrated rail factor)')
    _route_cache[key] = result
    return result


def fallback_distance_miles(origin, destination, mode):
    a = geocode_location(origin)
    b = geocode_location(destination)
    if a and b:
        gc = haversine_miles(a, b)
        mode_key = (mode or '').strip().lower()
        factor = {'road': ROAD_FACTOR, 'rail': RAIL_FACTOR, 'sea': SEA_FACTOR, 'inland water': INLAND_FACTOR}.get(mode_key, 1.25)
        miles = max(1, round(gc * factor))
        method = f'Estimated using great-circle × calibrated {mode_key or "transport"} factor'
        return miles, method
    mode_key = (mode or '').strip().lower()
    miles = {'road': 900, 'rail': 1400, 'sea': 3500, 'inland water': 600}.get(mode_key, 1200)
    method = f'Estimated using deterministic {mode_key or "transport"} fallback heuristic'
    return miles, method


def estimate_distance(origin, destination, mode):
    mode_key = (mode or '').strip().lower()
    cache_key = (normalize_key(origin), normalize_key(destination), mode_key)
    if cache_key in _distance_cache:
        return _distance_cache[cache_key]

    if mode_key == 'storage':
        result = (None, None)
    elif mode_key == 'road':
        road = road_distance_miles(origin, destination)
        if road is not None:
            result = (max(1, round(road)), 'Road estimate (online road network routing)')
        else:
            result = fallback_distance_miles(origin, destination, mode)
    elif mode_key == 'rail':
        rail = rail_distance_miles(origin, destination)
        result = rail if rail is not None else fallback_distance_miles(origin, destination, mode)
    elif mode_key == 'sea':
        sea = sea_distance_miles(origin, destination)
        result = sea if sea is not None else fallback_distance_miles(origin, destination, mode)
    elif mode_key == 'inland water':
        inland = inland_water_distance_miles(origin, destination)
        result = inland if inland is not None else fallback_distance_miles(origin, destination, mode)
    else:
        result = fallback_distance_miles(origin, destination, mode)

    _distance_cache[cache_key] = result
    return result


def compatible_formula(row_num, mode):
    r = row_num
    if mode == 'road':
        return f'=IF(F{r}="Road",((C{r}/2204.62)*M{r}*INDEX(\'Emission Factors\'!B:B,MATCH(IF(TRIM(J{r})="","Default road vehicle",TRIM(J{r})),\'Emission Factors\'!A:A,0)))/1000000,0)'
    if mode == 'rail':
        return f'=IF(F{r}="Rail",((C{r}/2204.62*M{r})*INDEX(\'Emission Factors\'!B:B,MATCH(IF(TRIM(H{r})="","Rail industry-average",TRIM(H{r})),\'Emission Factors\'!A:A,0)))/1000000,0)'
    if mode == 'inland':
        return f'=IF(F{r}="Inland Water",((C{r}/2204.62*M{r})*INDEX(\'Emission Factors\'!B:B,MATCH(IF(TRIM(L{r})="","Default inland water vehicle",TRIM(L{r})),\'Emission Factors\'!A:A,0)))/1000000,0)'
    if mode == 'sea':
        return f'=IF(F{r}="Sea",((C{r}/2204.62*M{r})*INDEX(\'Emission Factors\'!B:B,MATCH(IF(TRIM(K{r})="","Default sea vessel",TRIM(K{r})),\'Emission Factors\'!A:A,0)))/1000000,0)'
    if mode == 'storage':
        return f'=IF(F{r}="Storage",(C{r}/2204.62)*\'Emission Factors\'!B50/1000,0)'
    raise ValueError(mode)


def clear_row_a_to_n(ws, row_num):
    for c in COLS_A_TO_N:
        ws.cell(row_num, c).value = None


def process_workbooks(input_path, template_path, output_path=None, repair_formulas=True):
    input_wb = load_workbook(input_path, data_only=False)
    input_ws = input_wb[input_wb.sheetnames[0]]
    template_wb = load_workbook(template_path, data_only=False)
    output_ws = template_wb['Project Data and GHG Calcs']

    input_map = parse_headers(input_ws)
    missing_headers = [h for h in INPUT_HEADERS if h not in input_map]
    if missing_headers:
        raise ValueError(f'Input sheet is missing required headers: {", ".join(missing_headers)}')

    rows = []
    for row_num in range(2, input_ws.max_row + 1):
        row = {h: input_ws.cell(row_num, input_map[h]).value for h in INPUT_HEADERS}
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row.values()):
            continue
        weight = row['Weight (Pounds)']
        if isinstance(weight, str) and weight.strip().lower() == 'legal':
            row['Weight (Pounds)'] = 44000
        else:
            row['Weight (Pounds)'] = clean_output_value(weight)
        for key in INPUT_HEADERS:
            if key != 'Weight (Pounds)':
                row[key] = clean_output_value(row[key])
        if looks_blank_or_zero(row['Distance (Miles)']):
            estimate, method = estimate_distance(row['Origin'], row['Destination'], row['Mode'])
            row['Distance (Miles)'] = ensure_final_miles(estimate)
            row['Distance Estimation Method'] = method
        else:
            row['Distance (Miles)'] = ensure_final_miles(row['Distance (Miles)']) if row['Mode'] != 'Storage' else clean_output_value(row['Distance (Miles)'])
            row['Distance Estimation Method'] = None
        rows.append(row)

    def sort_key(item):
        value = item['Delivery Date']
        if isinstance(value, datetime):
            return value
        return datetime.max

    rows.sort(key=sort_key)

    start_row = 2
    for idx, row in enumerate(rows, start=start_row):
        output_ws.cell(idx, 1).value = row['Project ID']
        output_ws.cell(idx, 2).value = row['Pick Up Date']
        output_ws.cell(idx, 3).value = row['Weight (Pounds)']
        output_ws.cell(idx, 4).value = row['Origin']
        output_ws.cell(idx, 5).value = row['Destination']
        output_ws.cell(idx, 6).value = row['Mode']
        output_ws.cell(idx, 7).value = row['Delivery Date']
        output_ws.cell(idx, 8).value = row['Subcontractor/Partner']
        output_ws.cell(idx, 9).value = row['Client']
        output_ws.cell(idx, 10).value = row['(Road) Vehicle category (US)']
        output_ws.cell(idx, 11).value = row['(Sea) Vessel Type']
        output_ws.cell(idx, 12).value = row['(Inland Water) Vessel Type']
        output_ws.cell(idx, 13).value = row['Distance (Miles)']
        output_ws.cell(idx, 14).value = row['Distance Estimation Method']
        if repair_formulas:
            output_ws.cell(idx, 15).value = compatible_formula(idx, 'road')
            output_ws.cell(idx, 16).value = compatible_formula(idx, 'rail')
            output_ws.cell(idx, 17).value = compatible_formula(idx, 'inland')
            output_ws.cell(idx, 18).value = compatible_formula(idx, 'sea')
            output_ws.cell(idx, 19).value = compatible_formula(idx, 'storage')

    for row_num in range(start_row + len(rows), output_ws.max_row + 1):
        clear_row_a_to_n(output_ws, row_num)

    if output_path is None:
        stamp = datetime.now().strftime('%Y-%m-%d')
        output_path = str(Path(template_path).with_name(f'HLI Project GHG Calcs {stamp}.xlsx'))

    template_wb.save(output_path)
    return output_path, len(rows)


class App:
    def __init__(self, root):
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry('780x580')
        self.input_path = tk.StringVar()
        self.template_path = tk.StringVar()
        self.output_dir = tk.StringVar()
        self.repair_formulas = tk.BooleanVar(value=True)
        self.status = tk.StringVar(value='Select the input workbook and the template workbook.')
        self._build()

    def _build(self):
        frame = ttk.Frame(self.root, padding=16)
        frame.pack(fill='both', expand=True)

        ttk.Label(frame, text='HLI Project GHG Calcs Builder', font=('Segoe UI', 16, 'bold')).pack(anchor='w')
        ttk.Label(
            frame,
            text='Processes workbooks locally. No data is stored in any external database. This version uses online routing where available for road, maritime network logic for sea when the package is available, corridor logic for Atlantic ICW inland-water lanes, and stronger rail network-proxy logic.',
            wraplength=720,
        ).pack(anchor='w', pady=(6, 14))

        self._file_row(frame, 'Input workbook', self.input_path, self.pick_input)
        self._file_row(frame, 'Template workbook', self.template_path, self.pick_template)
        self._file_row(frame, 'Output folder', self.output_dir, self.pick_output_dir)

        ttk.Checkbutton(
            frame,
            text='Repair template formulas in O:S for Excel compatibility (recommended)',
            variable=self.repair_formulas,
        ).pack(anchor='w', pady=(10, 8))

        info = (
            'Rules applied:\n'
            '• Writes only columns A:N\n'
            '• Leaves the Emission Factors tab intact\n'
            '• Converts Legal weight to 44,000\n'
            '• Leaves 0, NAME?, and #NAME? blank in imported values\n'
            '• Sorts by Delivery Date, oldest first\n'
            '• Distance never remains blank except for Storage\n'
            '• Reuses cached lane estimates for faster processing\n'
            '• Road uses online road routing when available\n'
            '• Sea uses maritime network routing when available\n'
            '• Inland water prefers Atlantic ICW corridor logic when applicable\n'
            '• Rail uses corridor / hub routing before falling back to calibrated factors\n'
        )
        ttk.Label(frame, text=info, justify='left').pack(anchor='w', pady=(6, 10))

        btns = ttk.Frame(frame)
        btns.pack(fill='x', pady=(8, 10))
        self.process_btn = ttk.Button(btns, text='Process Workbooks', command=self.start_processing)
        self.process_btn.pack(side='left')
        ttk.Button(btns, text='Open Output Folder', command=self.open_output_folder).pack(side='left', padx=8)

        ttk.Label(frame, textvariable=self.status, foreground='#0a5').pack(anchor='w', pady=(6, 8))
        self.progress = ttk.Progressbar(frame, mode='indeterminate')
        self.progress.pack(fill='x')

        notes = (
            'Suggested QA lanes after rebuild:\n'
            '• Oakland, CA → Cancun, Mexico (Road)\n'
            '• Rotterdam, Netherlands → New York City, NY (Sea)\n'
            '• Myrtle Beach, SC → Miami, FL (Inland Water)\n'
            '• Boston, MA → Ithaca, NY (Rail)\n'
        )
        ttk.Label(frame, text=notes, justify='left').pack(anchor='w', pady=(12, 0))

    def _file_row(self, parent, label, variable, command):
        row = ttk.Frame(parent)
        row.pack(fill='x', pady=4)
        ttk.Label(row, text=label, width=18).pack(side='left')
        ttk.Entry(row, textvariable=variable).pack(side='left', fill='x', expand=True, padx=(0, 8))
        ttk.Button(row, text='Browse', command=command).pack(side='left')

    def pick_input(self):
        path = filedialog.askopenfilename(filetypes=[('Excel workbooks', '*.xlsx *.xlsm *.xltx *.xltm')])
        if path:
            self.input_path.set(path)

    def pick_template(self):
        path = filedialog.askopenfilename(filetypes=[('Excel workbooks', '*.xlsx *.xlsm *.xltx *.xltm')])
        if path:
            self.template_path.set(path)

    def pick_output_dir(self):
        path = filedialog.askdirectory()
        if path:
            self.output_dir.set(path)

    def open_output_folder(self):
        path = self.output_dir.get().strip()
        if path:
            Path(path).mkdir(parents=True, exist_ok=True)
            try:
                import os
                os.startfile(path)
            except Exception:
                messagebox.showinfo(APP_TITLE, f'Output folder: {path}')

    def start_processing(self):
        if not self.input_path.get().strip() or not self.template_path.get().strip():
            messagebox.showerror(APP_TITLE, 'Please select both the input workbook and the template workbook.')
            return
        self.process_btn.configure(state='disabled')
        self.progress.start(8)
        self.status.set('Processing...')
        worker = threading.Thread(target=self._process_thread, daemon=True)
        worker.start()

    def _process_thread(self):
        try:
            out_dir = Path(self.output_dir.get().strip()) if self.output_dir.get().strip() else Path(self.template_path.get()).parent
            out_dir.mkdir(parents=True, exist_ok=True)
            stamp = datetime.now().strftime('%Y-%m-%d')
            out_path = out_dir / f'HLI Project GHG Calcs {stamp}.xlsx'
            output_path, row_count = process_workbooks(
                self.input_path.get().strip(),
                self.template_path.get().strip(),
                str(out_path),
                repair_formulas=self.repair_formulas.get(),
            )
            self.root.after(0, lambda: self._on_success(output_path, row_count))
        except Exception as exc:
            self.root.after(0, lambda: self._on_error(exc))

    def _on_success(self, output_path, row_count):
        self.progress.stop()
        self.process_btn.configure(state='normal')
        self.status.set(f'Success: processed {row_count} rows → {output_path}')
        messagebox.showinfo(APP_TITLE, f'Successfully processed {row_count} rows.\n\nSaved to:\n{output_path}')

    def _on_error(self, exc):
        self.progress.stop()
        self.process_btn.configure(state='normal')
        self.status.set('Processing failed.')
        messagebox.showerror(APP_TITLE, f'Processing failed:\n\n{exc}')


def main():
    root = tk.Tk()
    style = ttk.Style()
    if 'vista' in style.theme_names():
        style.theme_use('vista')
    App(root)
    root.mainloop()


if __name__ == '__main__':
    main()
